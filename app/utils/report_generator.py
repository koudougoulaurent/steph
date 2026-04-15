"""
Générateur de rapports PDF et Excel
Utilise ReportLab pour PDF et openpyxl pour Excel
"""

import os
import json
from datetime import datetime, date
from flask import current_app
from sqlalchemy import func

from app import db
from app.models import (
    Couverture, ClasseCouverture, FeuxBrousse,
    SiteVulnerable, IndicateurBraconnage,
    ObservationTerrain, CampagneCollecte
)


class ReportGenerator:
    """
    Génère des rapports environnementaux structurés.
    Supporte PDF (ReportLab) et Excel (openpyxl).
    """

    def __init__(self, rapport):
        self.rapport = rapport
        self.debut = rapport.periode_debut
        self.fin = rapport.periode_fin
        self.output_dir = current_app.config['RAPPORTS_FOLDER']
        os.makedirs(self.output_dir, exist_ok=True)
        self._stats = None

    def _collect_stats(self):
        """Collecte toutes les statistiques de la période"""
        if self._stats:
            return self._stats

        debut = self.debut
        fin = self.fin

        # Feux de brousse
        feux = FeuxBrousse.query.filter(
            FeuxBrousse.date_debut >= debut,
            FeuxBrousse.date_debut <= fin
        ).all()
        superficie_feux = sum(f.superficie_brulee_ha or 0 for f in feux)

        # Sites vulnérables (actifs à la fin de la période)
        sites = SiteVulnerable.query.filter_by(statut='actif').all()
        sites_critiques = [s for s in sites if s.niveau_vulnerabilite == 'Critique']

        # Braconnage
        braconnage = IndicateurBraconnage.query.filter(
            IndicateurBraconnage.date_constat >= debut,
            IndicateurBraconnage.date_constat <= fin
        ).all()

        # Observations
        observations = ObservationTerrain.query.filter(
            ObservationTerrain.date_observation >= datetime.combine(debut, datetime.min.time()),
            ObservationTerrain.date_observation <= datetime.combine(fin, datetime.max.time())
        ).all()

        # Couverture végétale (années dans la période)
        annees_periode = list(range(debut.year, fin.year + 1))
        couverture_data = {}
        classes = ClasseCouverture.query.filter_by(actif=True).all()
        for annee in annees_periode:
            couverture_data[annee] = {}
            for classe in classes:
                total = db.session.query(
                    func.coalesce(func.sum(Couverture.superficie_ha), 0)
                ).filter_by(classe_id=classe.id, annee=annee).scalar()
                couverture_data[annee][classe.label] = round(float(total), 2)

        self._stats = {
            'feux': feux,
            'nb_feux': len(feux),
            'superficie_feux_ha': round(superficie_feux, 1),
            'sites': sites,
            'nb_sites': len(sites),
            'nb_sites_critiques': len(sites_critiques),
            'braconnage': braconnage,
            'nb_braconnage': len(braconnage),
            'observations': observations,
            'nb_observations': len(observations),
            'couverture_data': couverture_data,
            'classes': classes,
        }
        return self._stats

    # ──────────────── PDF ────────────────

    def generate_pdf(self) -> str:
        """Génère le rapport en PDF - retourne le chemin du fichier"""
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
            from reportlab.lib import colors
            from reportlab.platypus import (
                SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
                HRFlowable, PageBreak
            )
            from reportlab.lib.enums import TA_CENTER, TA_LEFT
        except ImportError:
            raise ImportError("ReportLab n'est pas installé. Exécutez : pip install reportlab")

        stats = self._collect_stats()
        filename = f"rapport_{self.rapport.reference}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        filepath = os.path.join(self.output_dir, filename)

        doc = SimpleDocTemplate(
            filepath,
            pagesize=A4,
            rightMargin=2 * cm,
            leftMargin=2 * cm,
            topMargin=2.5 * cm,
            bottomMargin=2 * cm
        )

        styles = getSampleStyleSheet()
        style_title = ParagraphStyle(
            'CustomTitle', parent=styles['Title'],
            fontSize=18, spaceAfter=6, textColor=colors.HexColor('#1a5276')
        )
        style_h2 = ParagraphStyle(
            'H2', parent=styles['Heading2'],
            fontSize=13, spaceBefore=12, spaceAfter=4,
            textColor=colors.HexColor('#1a5276'),
            borderPad=4
        )
        style_h3 = ParagraphStyle(
            'H3', parent=styles['Heading3'],
            fontSize=11, spaceBefore=8, spaceAfter=3,
            textColor=colors.HexColor('#117a65')
        )
        style_normal = styles['Normal']
        style_center = ParagraphStyle('Center', parent=styles['Normal'], alignment=TA_CENTER)

        GREEN = colors.HexColor('#1a5276')
        LIGHT_GREEN = colors.HexColor('#d5e8d4')
        HEADER_BG = colors.HexColor('#2e86c1')
        TABLE_HEADER = colors.HexColor('#1a5276')

        story = []

        # ── EN-TÊTE ──
        story.append(Paragraph('DIRECTION RÉGIONALE DE L\'ENVIRONNEMENT', style_center))
        story.append(Paragraph('VégéSuivi Pro — Système de Suivi de la Végétation', style_center))
        story.append(HRFlowable(width='100%', thickness=2, color=GREEN, spaceAfter=12))
        story.append(Paragraph(self.rapport.titre, style_title))
        story.append(Paragraph(
            f"Période : {self.debut.strftime('%d/%m/%Y')} au {self.fin.strftime('%d/%m/%Y')}",
            style_normal
        ))
        story.append(Paragraph(
            f"Zone : {self.rapport.zone_couverte or 'Toute la région'}   |   "
            f"Type : {self.rapport.get_type_label()}   |   "
            f"Réf. : {self.rapport.reference}",
            style_normal
        ))
        story.append(Paragraph(
            f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}",
            style_normal
        ))
        story.append(HRFlowable(width='100%', thickness=1, color=colors.grey, spaceAfter=16))

        # ── RÉSUMÉ EXÉCUTIF ──
        if self.rapport.resume_executif:
            story.append(Paragraph('I. RÉSUMÉ EXÉCUTIF', style_h2))
            story.append(Paragraph(self.rapport.resume_executif, style_normal))
            story.append(Spacer(1, 10))

        # ── KPI SYNTHÈSE ──
        story.append(Paragraph('II. INDICATEURS CLÉS DE LA PÉRIODE', style_h2))
        kpi_data = [
            ['Indicateur', 'Valeur', 'Observation'],
            ['Feux de brousse détectés', str(stats['nb_feux']),
             f"{stats['superficie_feux_ha']} ha brûlés"],
            ['Sites vulnérables actifs', str(stats['nb_sites']),
             f"dont {stats['nb_sites_critiques']} critiques"],
            ['Indicateurs de braconnage', str(stats['nb_braconnage']), 'Période couverte'],
            ['Observations terrain', str(stats['nb_observations']), 'Toutes catégories'],
        ]
        t = Table(kpi_data, colWidths=[7 * cm, 3 * cm, 7 * cm])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), TABLE_HEADER),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, LIGHT_GREEN]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ALIGN', (1, 0), (1, -1), 'CENTER'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]))
        story.append(t)
        story.append(Spacer(1, 14))

        # ── FEUX DE BROUSSE ──
        story.append(Paragraph('III. FEUX DE BROUSSE', style_h2))
        story.append(Paragraph(
            f"Total : <b>{stats['nb_feux']}</b> feux détectés — "
            f"Superficie totale brûlée : <b>{stats['superficie_feux_ha']} ha</b>",
            style_normal
        ))
        story.append(Spacer(1, 6))
        if stats['feux']:
            feux_data = [['Référence', 'Date', 'Zone', 'Superficie (ha)', 'Intensité', 'Cause']]
            for f in stats['feux'][:20]:
                feux_data.append([
                    f.reference or '-',
                    f.date_debut.strftime('%d/%m/%Y') if f.date_debut else '-',
                    (f.zone or '-')[:25],
                    str(f.superficie_brulee_ha or 0),
                    f.intensite or '-',
                    f.cause or '-',
                ])
            tf = Table(feux_data, colWidths=[3 * cm, 2.5 * cm, 4.5 * cm, 2.5 * cm, 2.5 * cm, 2 * cm])
            tf.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), TABLE_HEADER),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.3, colors.lightgrey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#fdf2e0')]),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ]))
            story.append(tf)
            if len(stats['feux']) > 20:
                story.append(Paragraph(
                    f"<i>... et {len(stats['feux']) - 20} autres enregistrements (voir export Excel).</i>",
                    style_normal))
        else:
            story.append(Paragraph('Aucun feu de brousse enregistré sur la période.', style_normal))
        story.append(Spacer(1, 12))

        # ── SITES VULNÉRABLES ──
        story.append(Paragraph('IV. SITES VULNÉRABLES', style_h2))
        if stats['sites']:
            sv_data = [['Référence', 'Nom', 'Type', 'Zone', 'Vulnérabilité', 'Score']]
            for s in sorted(stats['sites'],
                            key=lambda x: x.score_vulnerabilite or 0, reverse=True)[:15]:
                sv_data.append([
                    s.reference or '-',
                    (s.nom or '-')[:25],
                    (s.type_site or '-')[:18],
                    (s.zone or '-')[:18],
                    s.niveau_vulnerabilite or '-',
                    str(s.score_vulnerabilite or '-'),
                ])
            tsv = Table(sv_data, colWidths=[2.5 * cm, 4 * cm, 3 * cm, 3 * cm, 2.5 * cm, 1.5 * cm])
            tsv.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), TABLE_HEADER),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.3, colors.lightgrey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, LIGHT_GREEN]),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ]))
            story.append(tsv)
        story.append(Spacer(1, 12))

        # ── BRACONNAGE ──
        story.append(Paragraph('V. INDICATEURS DE BRACONNAGE', style_h2))
        if stats['braconnage']:
            br_data = [['Référence', 'Date', 'Type', 'Zone', 'Gravité', 'Statut']]
            for b in stats['braconnage'][:15]:
                br_data.append([
                    b.reference or '-',
                    b.date_constat.strftime('%d/%m/%Y') if b.date_constat else '-',
                    (b.type_indicateur or '-')[:20],
                    (b.zone or '-')[:20],
                    b.niveau_gravite or '-',
                    b.statut or '-',
                ])
            tbr = Table(br_data, colWidths=[2.5 * cm, 2.5 * cm, 4 * cm, 3.5 * cm, 2.5 * cm, 2.5 * cm])
            tbr.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), TABLE_HEADER),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.3, colors.lightgrey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#fce4e4')]),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ]))
            story.append(tbr)
        else:
            story.append(Paragraph('Aucun indicateur de braconnage sur la période.', style_normal))
        story.append(Spacer(1, 12))

        # ── COUVERTURE VEGÉTALE ──
        story.append(Paragraph('VI. COUVERTURE VÉGÉTALE', style_h2))
        for annee, data in stats['couverture_data'].items():
            if data:
                story.append(Paragraph(f'Année {annee}', style_h3))
                vc_data = [['Classe d\'occupation', 'Superficie (ha)', 'Superficie (km²)']]
                for label, sup_ha in data.items():
                    vc_data.append([label, f"{sup_ha:.1f}", f"{sup_ha/100:.2f}"])
                tvc = Table(vc_data, colWidths=[9 * cm, 4 * cm, 4 * cm])
                tvc.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#117a65')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('GRID', (0, 0), (-1, -1), 0.3, colors.lightgrey),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, LIGHT_GREEN]),
                    ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
                    ('TOPPADDING', (0, 0), (-1, -1), 4),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ]))
                story.append(tvc)
                story.append(Spacer(1, 8))

        # ── PIED DE PAGE ──
        story.append(PageBreak())
        story.append(HRFlowable(width='100%', thickness=1, color=GREEN))
        story.append(Paragraph(
            f"<i>VégéSuivi Pro — Direction Régionale de l'Environnement — "
            f"Document généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}</i>",
            style_center
        ))

        doc.build(story)
        return filepath

    # ──────────────── EXCEL ────────────────

    def generate_excel(self) -> str:
        """Génère le rapport en Excel - retourne le chemin du fichier"""
        try:
            import openpyxl
            from openpyxl.styles import (
                Font, PatternFill, Alignment, Border, Side
            )
            from openpyxl.chart import BarChart, Reference, LineChart
        except ImportError:
            raise ImportError("openpyxl n'est pas installé. Exécutez : pip install openpyxl")

        stats = self._collect_stats()
        filename = f"rapport_{self.rapport.reference}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(self.output_dir, filename)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Enlève la feuille vide par défaut

        # Couleurs
        C_GREEN = '1A5276'
        C_LIGHT = 'D5E8D4'
        C_ORANGE = 'F0B27A'
        C_RED = 'F1948A'
        C_WHITE = 'FFFFFF'

        def header_style(cell, bg=C_GREEN):
            cell.font = Font(bold=True, color=C_WHITE, size=11)
            cell.fill = PatternFill('solid', fgColor=bg)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        def add_sheet_header(ws, titre, ref, periode):
            ws.merge_cells('A1:G1')
            ws['A1'] = 'DIRECTION RÉGIONALE DE L\'ENVIRONNEMENT — VégéSuivi Pro'
            ws['A1'].font = Font(bold=True, size=13, color=C_GREEN)
            ws['A1'].alignment = Alignment(horizontal='center')
            ws.merge_cells('A2:G2')
            ws['A2'] = titre
            ws['A2'].font = Font(bold=True, size=11)
            ws['A2'].alignment = Alignment(horizontal='center')
            ws.merge_cells('A3:G3')
            ws['A3'] = f"Réf: {ref} | Période: {periode}"
            ws['A3'].alignment = Alignment(horizontal='center')
            ws['A3'].font = Font(italic=True)

        # ── Feuille RÉSUMÉ ──
        ws = wb.create_sheet('Résumé')
        add_sheet_header(ws, self.rapport.titre, self.rapport.reference,
                         f"{self.debut.strftime('%d/%m/%Y')} - {self.fin.strftime('%d/%m/%Y')}")
        ws.row_dimensions[5].height = 20
        headers = ['Indicateur', 'Valeur']
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=5, column=col, value=h)
            header_style(c)
        data_summary = [
            ('Feux de brousse détectés', stats['nb_feux']),
            ('Superficie brûlée (ha)', stats['superficie_feux_ha']),
            ('Sites vulnérables actifs', stats['nb_sites']),
            ('Sites vulnérables critiques', stats['nb_sites_critiques']),
            ('Indicateurs de braconnage', stats['nb_braconnage']),
            ('Observations terrain', stats['nb_observations']),
        ]
        for i, (label, val) in enumerate(data_summary, 6):
            ws.cell(row=i, column=1, value=label)
            ws.cell(row=i, column=2, value=val)
            if i % 2 == 0:
                for col in range(1, 3):
                    ws.cell(row=i, column=col).fill = PatternFill('solid', fgColor=C_LIGHT)
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 15

        # ── Feuille FEUX ──
        ws_f = wb.create_sheet('Feux de brousse')
        add_sheet_header(ws_f, 'Feux de brousse', self.rapport.reference, '')
        headers_f = ['Référence', 'Date début', 'Zone', 'Village', 'Superficie (ha)',
                     'Intensité', 'Cause', 'Impact faune', 'Statut']
        for col, h in enumerate(headers_f, 1):
            c = ws_f.cell(row=5, column=col, value=h)
            header_style(c)
        for i, f in enumerate(stats['feux'], 6):
            row_data = [
                f.reference, str(f.date_debut), f.zone, f.village_proche,
                f.superficie_brulee_ha, f.intensite, f.cause, f.impact_faune, f.statut
            ]
            for col, val in enumerate(row_data, 1):
                ws_f.cell(row=i, column=col, value=val)
        for col in range(1, len(headers_f) + 1):
            ws_f.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 16

        # ── Feuille BRACONNAGE ──
        ws_b = wb.create_sheet('Braconnage')
        add_sheet_header(ws_b, 'Indicateurs de braconnage', self.rapport.reference, '')
        headers_b = ['Référence', 'Date constat', 'Type', 'Zone', 'Gravité',
                     'Nb indices', 'Saisies', 'Arrestations', 'Statut']
        for col, h in enumerate(headers_b, 1):
            c = ws_b.cell(row=5, column=col, value=h)
            header_style(c, C_RED[:6] if len(C_RED) == 6 else C_GREEN)
        for i, b in enumerate(stats['braconnage'], 6):
            row_data = [
                b.reference, str(b.date_constat), b.type_indicateur, b.zone,
                b.niveau_gravite, b.nombre_indices, b.saisies_effectuees,
                b.arrestations, b.statut
            ]
            for col, val in enumerate(row_data, 1):
                ws_b.cell(row=i, column=col, value=val)
        for col in range(1, len(headers_b) + 1):
            ws_b.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 16

        # ── Feuille SITES VULNÉRABLES ──
        ws_sv = wb.create_sheet('Sites vulnérables')
        add_sheet_header(ws_sv, 'Sites vulnérables', self.rapport.reference, '')
        headers_sv = ['Référence', 'Nom', 'Type', 'Zone', 'Vulnérabilité', 'Score',
                      'Superficie (ha)', 'Fréquence surveillance', 'Statut']
        for col, h in enumerate(headers_sv, 1):
            c = ws_sv.cell(row=5, column=col, value=h)
            header_style(c, 'C0392B')
        for i, s in enumerate(stats['sites'], 6):
            row_data = [
                s.reference, s.nom, s.type_site, s.zone,
                s.niveau_vulnerabilite, s.score_vulnerabilite,
                s.superficie_ha, s.frequence_surveillance, s.statut
            ]
            for col, val in enumerate(row_data, 1):
                ws_sv.cell(row=i, column=col, value=val)
        for col in range(1, len(headers_sv) + 1):
            ws_sv.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

        # ── Feuille COUVERTURE VÉGÉTALE ──
        ws_cv = wb.create_sheet('Couverture végétale')
        add_sheet_header(ws_cv, 'Données de couverture végétale', self.rapport.reference, '')
        row = 5
        for annee, data in stats['couverture_data'].items():
            ws_cv.cell(row=row, column=1, value=f'Année {annee}')
            ws_cv.cell(row=row, column=1).font = Font(bold=True, size=12, color=C_GREEN)
            row += 1
            for col, h in enumerate(['Classe d\'occupation', 'Superficie (ha)', 'Superficie (km²)'], 1):
                c = ws_cv.cell(row=row, column=col, value=h)
                header_style(c, '117A65')
            row += 1
            for label, sup_ha in data.items():
                ws_cv.cell(row=row, column=1, value=label)
                ws_cv.cell(row=row, column=2, value=round(sup_ha, 2))
                ws_cv.cell(row=row, column=3, value=round(sup_ha / 100, 4))
                row += 1
            row += 1
        ws_cv.column_dimensions['A'].width = 35
        ws_cv.column_dimensions['B'].width = 18
        ws_cv.column_dimensions['C'].width = 18

        # ── Feuille OBSERVATIONS ──
        ws_obs = wb.create_sheet('Observations terrain')
        add_sheet_header(ws_obs, 'Observations terrain', self.rapport.reference, '')
        headers_obs = ['Référence', 'Date', 'Catégorie', 'Zone', 'Titre',
                       'État', 'Alerte', 'Validée', 'Agent']
        for col, h in enumerate(headers_obs, 1):
            c = ws_obs.cell(row=5, column=col, value=h)
            header_style(c, '1F618D')
        for i, o in enumerate(stats['observations'], 6):
            agent_nom = o.agent.nom if o.agent else '-'
            row_data = [
                o.reference, str(o.date_observation.date()), o.categorie, o.zone,
                o.titre, o.etat_general, o.niveau_alerte,
                'Oui' if o.validee else 'Non', agent_nom
            ]
            for col, val in enumerate(row_data, 1):
                ws_obs.cell(row=i, column=col, value=val)
        for col in range(1, len(headers_obs) + 1):
            ws_obs.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 16

        wb.save(filepath)
        return filepath
